import numpy as np
from numpy import *
import openpyxl as ol
from openpyxl import *
from numba import njit, jit
import os
from os import *
import sys
from PyQt5 import QtGui
from PyQt5.QtWidgets import *   # pip install pyqt5 , библиотека для создания интерфейса
from PyQt5.QtWebEngineWidgets import *      # pip install PyQtWebEngine
from PyQt5.QtCore import *

is_static_build = True

if is_static_build:
    xl_path='E:/ВУЗ/Проект/Транзит 2019-2020 гг..xlsx'
else:
    xl_path=input('Введите полный путь до Excel-таблицы (вместе с расширением) ')

xl_file = load_workbook(filename=xl_path,  data_only=True)  # Создание мелкой копии файла. Атрибут data_only позволяет избежать получения формул при попытке получить данные из ячейки
working_sheet=xl_file[xl_file.sheetnames[0]]    # Выбор таблицы. xl_file.sheetnames возвращает список с листами таблицы. Указывая [0] мы передаем проге инфу название нужного листа и он полнстью записывается в working_sheet
max_rows=working_sheet.max_row
print('Sheet loaded')

#test_array = np.empty(1)
test_array=['']
values_dict = {}

#@jit
def sheet_dig():
	#sheet = load_workbook(filename='E:/ВУЗ/Проект/Транзит 2019-2020 гг..xlsx',  data_only=True)[load_workbook(filename='E:/ВУЗ/Проект/Транзит 2019-2020 гг..xlsx',  data_only=True).sheetnames[0]]
	for row_number in range(2, max_rows+1):
		for column_number in range(1, 16+1):
			add = str(working_sheet.cell(row=row_number, column=column_number).value)
			#print(add)
			test_array.append(add)

def column_dicts():
	

	for column_number in range(1, working_sheet.max_column+1):
		if working_sheet.cell(row=1, column=column_number).value == None:
			break
		else:
			column_values=[]
			for row_number in range(2, working_sheet.max_row+1):
				value=working_sheet.cell(row=row_number, column=column_number).value

				if value == None:
					break

				else:
					if value not in column_values:
						if isinstance(value, float):
							value=round(value, 2)
							column_values.append(value)
			column_values.sort()
			values_dict.update({working_sheet.cell(row=1, column=column_number).value : column_values})


#sheet_dig()
column_dicts()

print(values_dict)


app = QApplication(sys.argv)

window = QWidget()
window.show()


sys.exit(app.exec_())